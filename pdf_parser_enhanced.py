"""
AP2 PDF Parser - Enhanced with Advanced Table Extraction for Annual Reports
Uses PyMuPDF (fitz) and Camelot for accurate table extraction
Extracts various financial data points from annual reports
"""

import os
import glob
import pandas as pd
import fitz  # PyMuPDF
import camelot
from datetime import datetime
import re
import logging

import config
# from llm_extractor import extract_annual_data_llm # Will be created later

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

print("=" * 80)
print("AP2 PDF Parser - Enhanced Table Extraction Version (Annual Reports)")
print("=" * 80)


def find_latest_download_folder():
    """Find the most recent download folder"""
    download_folders = glob.glob(os.path.join(config.DOWNLOADS_DIR, '*'))
    if not download_folders:
        logger.error("No download folders found")
        return None

    latest_folder = max(download_folders, key=os.path.getmtime)
    logger.info(f"Processing folder: {latest_folder}")
    return latest_folder


def extract_year_from_filename(filename):
    """Extract year from various filename patterns"""
    basename = os.path.basename(filename)

    patterns = [
        r'(\d{4})',  # Any 4 digits
        r'AP2_(\d{4})',  # AP2_YYYY
        r'Annual.*?(\d{4})',  # Annual-Report-YYYY
    ]

    for pattern in patterns:
        match = re.search(pattern, basename)
        if match:
            year = int(match.group(1))
            if 2000 <= year <= 2030:
                return year

    return datetime.now().year


def clean_number_string(value_str, allow_decimal=False, is_percentage=False):
    """Clean and convert Swedish formatted numbers
    Handles: '184 676', '-2 410', '464970', '458.0', '-2.4', '9%', etc.

    Args:
        value_str: String containing the number
        allow_decimal: If True, returns float for decimal numbers, else int
        is_percentage: If True, removes '%' and converts to float
    """
    if pd.isna(value_str) or value_str is None:
        return None

    # Convert to string and clean
    value_str = str(value_str).strip()

    if not value_str or value_str == '' or value_str == '-':
        return None

    if is_percentage:
        value_str = value_str.replace('%', '').strip()

    # Remove spaces and non-breaking spaces, keep decimal points
    cleaned = value_str.replace(' ', '').replace(',', '').replace('\xa0', '')

    # Check if it has a decimal point
    has_decimal = '.' in cleaned

    try:
        if has_decimal or allow_decimal or is_percentage:
            return float(cleaned)
        else:
            return int(cleaned)
    except ValueError:
        return None


def find_page_with_keywords(pdf_path, keywords, min_score=1):
    """Find page containing specific keywords"""
    doc = fitz.open(pdf_path)
    best_page = None
    best_score = 0

    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text().lower()

        score = sum(1 for keyword in keywords if keyword in text)

        if score > best_score:
            best_score = score
            best_page = page_num + 1  # 1-indexed

    doc.close()

    if best_page and best_score >= min_score:
        return best_page

    return None


def extract_table_with_camelot(pdf_path, page_num, flavor='lattice'):
    """Extract table using Camelot with specified flavor"""
    logger.info(f"  Attempting Camelot ({flavor}) extraction on page {page_num}...")

    try:
        tables = camelot.read_pdf(
            pdf_path,
            pages=str(page_num),
            flavor=flavor,
            line_scale=40,
            strip_text='\n'
        )

        if len(tables) == 0:
            logger.warning(f"  Camelot ({flavor}) found no tables")
            return None

        logger.info(f"  Camelot ({flavor}) found {len(tables)} table(s)")
        return tables[0].df # Return the first detected table for now

    except Exception as e:
        logger.error(f"  Camelot ({flavor}) extraction failed: {e}")
        return None


def extract_fund_capital_flows_results(pdf_path, year):
    """Extract Fund capital, flows and results from the 'five-year overview' table"""
    logger.info(f"\n  Extracting 'Fund capital, flows and results' for {year}...")

    # Keywords to find the relevant page
    keywords = ['fund capital, flows and results', 'five-year overview', 'net outflows', 'net result']
    page_num = find_page_with_keywords(pdf_path, keywords, min_score=2)

    if not page_num:
        logger.warning("  Could not find 'Fund capital, flows and results' page.")
        return {}

    df = extract_table_with_camelot(pdf_path, page_num, flavor='lattice')
    if df is None:
        df = extract_table_with_camelot(pdf_path, page_num, flavor='stream')

    if df is None:
        logger.error("  Failed to extract table for 'Fund capital, flows and results'.")
        return {}

    data = {}
    # Find the column for the target year
    target_col_idx = -1
    for col_idx, col_val in enumerate(df.iloc[0]): # Assuming year is in the first row 
        if str(col_val).strip() == str(year):
            target_col_idx = col_idx
            break
    
    if target_col_idx == -1:
        logger.warning(f"  Could not find column for year {year} in 'Fund capital, flows and results' table.")
        return {}

    for idx, row in df.iterrows():
        field_name_raw = str(row.iloc[0]).strip().lower() if not pd.isna(row.iloc[0]) else ""
        value = clean_number_string(row.iloc[target_col_idx])

        if value is None:
            continue

        if 'fund capital' in field_name_raw and 'carried forward' in field_name_raw:
            data['FUNDCAPITALCARRIEDFORWARDLEVEL'] = value
        elif 'net outflows to the national pension system' in field_name_raw:
            data['NETOUTFLOWSTOTHENATIONALPENSIONSYSTEM'] = value
        elif 'net result for the year' in field_name_raw:
            data['TOTAL'] = value
    
    logger.info(f"  Extracted {len(data)} fields from 'Fund capital, flows and results'.")
    return data


def extract_asset_class_exposure(pdf_path, year):
    """Extract Asset class exposure data"""
    logger.info(f"\n  Extracting 'Asset class exposure' for {year}...")

    keywords = ['asset class exposure', 'strategic asset allocation', 'actual portfolio exposure']
    page_num = find_page_with_keywords(pdf_path, keywords, min_score=2)

    if not page_num:
        logger.warning("  Could not find 'Asset class exposure' page.")
        return {}

    df = extract_table_with_camelot(pdf_path, page_num, flavor='lattice')
    if df is None:
        df = extract_table_with_camelot(pdf_path, page_num, flavor='stream')

    if df is None:
        logger.error("  Failed to extract table for 'Asset class exposure'.")
        return {}

    data = {}
    # This table is complex, with multiple columns for percentages and SEK billion.
    # We need to identify the correct columns for 'Strategic asset allocation, %', 'Actual portfolio exposure, %', and 'SEK billion' for the target year.
    # For simplicity, let's assume we are looking for the latest year's data in the table.
    # The runbook indicates "Strategic Portfolio, %" and "%" (2nd column) and "SEK billion"
    
    # For now, a simplified approach: iterate through rows and try to match keywords
    # This part will likely need significant refinement based on actual table structure
    
    # Example: df.iloc[0] might contain headers like 'Strategic asset allocation, %', 'Actual portfolio exposure, %', 'SEK billion'
    
    # Let's try to find the column for SEK billion and the percentage columns
    sek_billion_col_idx = -1
    actual_allocation_percent_col_idx = -1
    
    # Assuming headers are in the first row or second row
    header_row_idx = 0
    if 'SEK billion' not in df.iloc[header_row_idx].astype(str).values:
        header_row_idx = 1 # Try next row if not found in first

    for col_idx, col_val in enumerate(df.iloc[header_row_idx]):
        col_val_str = str(col_val).strip().lower()
        if 'sek billion' in col_val_str:
            sek_billion_col_idx = col_idx
        if 'actual portfolio exposure' in col_val_str and '%' in col_val_str:
            actual_allocation_percent_col_idx = col_idx

    if sek_billion_col_idx == -1 and actual_allocation_percent_col_idx == -1:
        logger.warning("  Could not find 'SEK billion' or 'Actual portfolio exposure %' columns in 'Asset class exposure' table.")
        return {}

    # Iterate through rows to extract data
    for idx, row in df.iterrows():
        if idx <= header_row_idx: # Skip header rows
            continue

        field_name_raw = str(row.iloc[0]).strip().lower() if not pd.isna(row.iloc[0]) else ""
        
        # Map field names to config.OUTPUT_HEADERS
        # This is a simplified mapping and needs to be more robust
        
        # Example:
        if 'swedish equities' in field_name_raw:
            if actual_allocation_percent_col_idx != -1:
                data['DOMESTICEQUITIES.ACTUALALLOCATION.STRATEGICPORTFOLIO.NONE.A.1@AP2'] = clean_number_string(row.iloc[actual_allocation_percent_col_idx], is_percentage=True)
            if sek_billion_col_idx != -1:
                # This is a placeholder, the actual header for SEK billion for domestic equities needs to be determined
                pass 
        # ... similar logic for other fields based on the runbook and config.py
        
    logger.info(f"  Extracted {len(data)} fields from 'Asset class exposure'.")
    return data


def extract_real_assets_distribution(pdf_path, year):
    """Extract Real assets portfolio and geographical distribution"""
    logger.info(f"\n  Extracting 'Real assets distribution' for {year}...")

    keywords = ['real assets', 'portfolio distribution', 'geographical distribution']
    page_num = find_page_with_keywords(pdf_path, keywords, min_score=2)

    if not page_num:
        logger.warning("  Could not find 'Real assets distribution' page.")
        return {}

    # This section often contains pie charts, so direct table extraction might be tricky.
    # We might need to rely on text extraction and regex or LLM for this.
    # For now, let's try to extract any tables and then process text.

    doc = fitz.open(pdf_path)
    page = doc[page_num - 1]
    text = page.get_text()
    doc.close()

    data = {}

    # Example: Extracting from text using regex (placeholder)
    # This will need to be refined based on actual text patterns
    
    # Portfolio distribution
    # Traditional real estate
    match = re.search(r'Traditional real estate\s+(\d+)%', text)
    if match:
        data['TRADITIONALREALESTATE.ACTUALALLOCATION.NONE.A.1@AP2'] = float(match.group(1))

    # Natural Climate Solutions
    match = re.search(r'Natural Climate Solutions\s+(\d+)%', text)
    if match:
        data['NATURALCLIMATE.ACTUALALLOCATION.NONE.A.1@AP2'] = float(match.group(1))

    # Sustainable infrastructure
    match = re.search(r'Sustainable infrastructure\s+(\d+)%', text)
    if match:
        data['SUSTAINABLEINFRASTRUCTURE.ACTUALALLOCATION.NONE.A.1@AP2'] = float(match.group(1))

    # Geographical distribution
    # North America
    match = re.search(r'North America\s+(\d+)%', text)
    if match:
        data['NORTHAMERICA.ACTUALALLOCATION.NONE.A.1@AP2'] = float(match.group(1))
    
    # South America
    match = re.search(r'South America\s+(\d+)%', text)
    if match:
        data['SOUTHAMERICA.ACTUALALLOCATION.NONE.A.1@AP2'] = float(match.group(1))

    # Oceania
    match = re.search(r'Oceania\s+(\d+)%', text)
    if match:
        data['OCEANIA.ACTUALALLOCATION.NONE.A.1@AP2'] = float(match.group(1))

    # Europe (excl. Sweden)
    match = re.search(r'Europe \(excl\. Sweden\)\s+(\d+)%', text)
    if match:
        data['EUROPEEXCLSWEDEN.ACTUALALLOCATION.NONE.A.1@AP2'] = float(match.group(1))

    # Sweden
    match = re.search(r'Sweden\s+(\d+)%', text)
    if match:
        data['SWEDEN.ACTUALALLOCATION.NONE.A.1@AP2'] = float(match.group(1))

    # Asia
    match = re.search(r'Asia\s+(\d+)%', text)
    if match:
        data['ASIA.ACTUALALLOCATION.NONE.A.1@AP2'] = float(match.group(1))

    # Africa
    match = re.search(r'Africa\s+(\d+)%', text)
    if match:
        data['AFRICA.ACTUALALLOCATION.NONE.A.1@AP2'] = float(match.group(1))

    logger.info(f"  Extracted {len(data)} fields from 'Real assets distribution'.")
    return data


def extract_bonds_fixed_income(pdf_path, year):
    """Extract Bonds and other fixed-income securities data"""
    logger.info(f"\n  Extracting 'Bonds and other fixed-income securities' for {year}...")

    keywords = ['bonds and other fixed-income securities', 'breakdown by issuer category', 'breakdown by type of instrument']
    page_num = find_page_with_keywords(pdf_path, keywords, min_score=2)

    if not page_num:
        logger.warning("  Could not find 'Bonds and other fixed-income securities' page.")
        return {}

    df = extract_table_with_camelot(pdf_path, page_num, flavor='lattice')
    if df is None:
        df = extract_table_with_camelot(pdf_path, page_num, flavor='stream')

    if df is None:
        logger.error("  Failed to extract table for 'Bonds and other fixed-income securities'.")
        return {}

    data = {}
    # This table has two years of data (31 Dec 2024 and 31 Dec 2023)
    # We need to extract both. The runbook says "Collect values for both years".
    
    # Find column indices for 31 Dec 2024 and 31 Dec 2023
    col_2024_idx = -1
    col_2023_idx = -1

    # Assuming headers are in the first row
    for col_idx, col_val in enumerate(df.iloc[0]): # Assuming headers are in the first row
        col_val_str = str(col_val).strip()
        if '31 Dec 2024' in col_val_str:
            col_2024_idx = col_idx
        elif '31 Dec 2023' in col_val_str:
            col_2023_idx = col_idx

    if col_2024_idx == -1 or col_2023_idx == -1:
        logger.warning("  Could not find both 2024 and 2023 columns in 'Bonds and other fixed-income securities' table.")
        return {}

    # Iterate through rows to extract data for both years
    for idx, row in df.iterrows():
        if idx == 0: # Skip header row
            continue

        field_name_raw = str(row.iloc[0]).strip().lower() if not pd.isna(row.iloc[0]) else ""
        
        # This mapping needs to be precise based on the config.OUTPUT_HEADERS
        # For now, a simplified approach for the fields present in the config
        
        # Example:
        if 'government bonds in developed markets' in field_name_raw:
            data[f'GOVBONDSDEVMARKETS.ACTUALALLOCATION.STRATEGICPORTFOLIO.NONE.A.1@AP2_{year}'] = clean_number_string(row.iloc[col_2024_idx])
            data[f'GOVBONDSDEVMARKETS.ACTUALALLOCATION.STRATEGICPORTFOLIO.NONE.A.1@AP2_{year-1}'] = clean_number_string(row.iloc[col_2023_idx])
        # ... similar logic for other fields
        
    logger.info(f"  Extracted {len(data)} fields from 'Bonds and other fixed-income securities'.")
    return data


def extract_balance_sheet_annual(pdf_path, year):
    """Extract Balance Sheet data for annual reports"""
    logger.info(f"\n  Extracting 'Balance sheet' for {year}...")

    keywords = ['balance sheet', 'assets', 'liabilities', 'fund capital']
    page_num = find_page_with_keywords(pdf_path, keywords, min_score=2)

    if not page_num:
        logger.warning("  Could not find 'Balance sheet' page.")
        return {}

    df = extract_table_with_camelot(pdf_path, page_num, flavor='lattice')
    if df is None:
        df = extract_table_with_camelot(pdf_path, page_num, flavor='stream')

    if df is None:
        logger.error("  Failed to extract table for 'Balance sheet'.")
        return {}

    data = {}
    # The runbook shows columns for "30 June 2024", "30 June 2023", "31 Dec. 2023".
    # For annual reports, we are interested in the latest year-end data.
    # So for year 2023, we would take "31 Dec. 2023" column.
    # For year 2024, we would look for "31 Dec. 2024" (if available in the report).
    
    # Find the column for the latest year-end data
    target_col_idx = -1
    for col_idx, col_val in enumerate(df.iloc[0]): # Assuming headers are in the first row
        col_val_str = str(col_val).strip()
        if f'31 Dec. {year}' in col_val_str:
            target_col_idx = col_idx
            break
        elif f'31 Dec. {year-1}' in col_val_str: # Fallback for previous year if current year not found
            target_col_idx = col_idx
            year = year - 1 # Adjust year to match extracted data
            break

    if target_col_idx == -1:
        logger.warning(f"  Could not find year-end column for {year} or {year-1} in 'Balance sheet' table.")
        return {}

    # Iterate through rows to extract data
    for idx, row in df.iterrows():
        if idx == 0: # Skip header row
            continue

        field_name_raw = str(row.iloc[0]).strip().lower() if not pd.isna(row.iloc[0]) else ""
        value = clean_number_string(row.iloc[target_col_idx])

        if value is None:
            continue

        # This mapping needs to be precise based on the config.OUTPUT_HEADERS
        # Example:
        if 'equities and participations' in field_name_raw and 'listed' in field_name_raw:
            data['EQUITIESANDPARTICIPATIONSLISTED'] = value
        # ... similar logic for other balance sheet fields
        
    logger.info(f"  Extracted {len(data)} fields from 'Balance sheet'.")
    return data


def parse_annual_report_adaptive(pdf_path, year):
    """Parse annual report using multiple extraction methods"""
    logger.info(f"\nProcessing: {os.path.basename(pdf_path)} (Year: {year})")

    all_extracted_data = {}

    # Extract Fund capital, flows and results
    all_extracted_data.update(extract_fund_capital_flows_results(pdf_path, year))

    # Extract Asset class exposure
    all_extracted_data.update(extract_asset_class_exposure(pdf_path, year))

    # Extract Real assets distribution
    all_extracted_data.update(extract_real_assets_distribution(pdf_path, year))

    # Extract Bonds and other fixed-income securities
    all_extracted_data.update(extract_bonds_fixed_income(pdf_path, year))

    # Extract Balance Sheet
    all_extracted_data.update(extract_balance_sheet_annual(pdf_path, year))

    # LLM fallback (will be implemented later)
    # if len(all_extracted_data) < expected_total_fields: # Placeholder for total fields
    #     logger.info("  Trying LLM fallback for missing fields...")
    #     llm_data = extract_annual_data_llm(pdf_path, year)
    #     for key, value in llm_data.items():
    #         if key not in all_extracted_data:
    #             all_extracted_data[key] = value

    logger.info(f"\n  [FINAL] Extracted {len(all_extracted_data)} total fields for {year}.")

    return all_extracted_data


def create_output(all_data):
    """Create Excel output matching sample structure"""
    logger.info(f"\n{'='*80}")
    logger.info("CREATING OUTPUT")
    logger.info(f"{ '='*80}")

    # Create DataFrame
    df_data = []

    # Sort years for consistent output order
    sorted_years = sorted(all_data.keys())

    for year in sorted_years:
        data = all_data[year]
        row = {'Unnamed: 0': year}

        # Map extracted fields to output headers
        for header in config.OUTPUT_HEADERS[1:]:
            value = None

            # Extract field name from header (e.g., 'AP2.FUNDCAPITALCARRIEDFORWARD.LEVEL.NONE.A.1@AP2' -> 'FUNDCAPITALCARRIEDFORWARDLEVEL')
            parts = header.split('.')
            field_name_base = parts[1] # e.g., FUNDCAPITALCARRIEDFORWARD

            # Handle special cases for LEVEL, FLOW, etc.
            if len(parts) >= 3 and parts[2] == 'LEVEL':
                field_name = field_name_base + 'LEVEL'
            elif len(parts) >= 3 and parts[2] == 'FLOW':
                field_name = field_name_base # FLOW is often implied by the base name
            else:
                field_name = field_name_base

            # Check for year-specific keys (e.g., for bonds table)
            if f'{field_name}_{year}' in data:
                value = data[f'{field_name}_{year}']
            elif field_name in data:
                value = data[field_name]

            row[header] = value

        df_data.append(row)

    df = pd.DataFrame(df_data, columns=config.OUTPUT_HEADERS)

    # Create output folders
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_folder = os.path.join(config.OUTPUT_DIR, timestamp)
    latest_folder = os.path.join(config.OUTPUT_DIR, config.LATEST_FOLDER_NAME)

    os.makedirs(output_folder, exist_ok=True)
    os.makedirs(latest_folder, exist_ok=True)

    # Save files with 2 header rows (technical + sub-headers)
    output_file = os.path.join(output_folder, f'AP2_Annual_Financial_Data_{timestamp}.xlsx')
    latest_file = os.path.join(latest_folder, 'AP2_Annual_Financial_Data_latest.xlsx')

    # Save with custom header structure using openpyxl
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows

    for file_path in [output_file, latest_file]:
        wb = openpyxl.Workbook()
        ws = wb.active

        # Row 1: Technical headers (first column should be None, not 'Unnamed: 0')
        for col_idx, header in enumerate(config.OUTPUT_HEADERS, start=1):
            value = None if header == 'Unnamed: 0' else header
            ws.cell(row=1, column=col_idx, value=value)

        # Row 2: Human-readable sub-headers
        for col_idx, subheader in enumerate(config.OUTPUT_SUBHEADERS, start=1):
            ws.cell(row=2, column=col_idx, value=subheader)

        # Row 3+: Data
        for row_idx, (_, row_data) in enumerate(df.iterrows(), start=3):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(file_path)

    logger.info(f"✓ Saved: {output_file}")
    logger.info(f"✓ Saved: {latest_file}")

    # Calculate accuracy
    # Exclude the first column ('Unnamed: 0') from accuracy calculation
    filled_count = df.drop(columns=['Unnamed: 0']).notna().sum().sum()
    total_cells = len(df) * (len(config.OUTPUT_HEADERS) - 1)
    accuracy = (filled_count/total_cells*100) if total_cells > 0 else 0

    logger.info(f"✓ Filled {filled_count}/{total_cells} cells ({accuracy:.1f}%)")

    if accuracy >= 90:
        logger.info("[EXCELLENT] High extraction accuracy!")
    elif accuracy >= 70:
        logger.info("[GOOD] Acceptable extraction accuracy")
    else:
        logger.warning("[POOR] Low extraction accuracy")

    return output_file


def main():
    """Main execution"""
    try:
        # Find latest download folder
        download_folder = find_latest_download_folder()
        if not download_folder:
            return

        # Find PDFs
        pdf_files = glob.glob(os.path.join(download_folder, '*.pdf'))
        logger.info(f"Found {len(pdf_files)} PDF file(s)\n")

        if not pdf_files:
            logger.error("No PDF files found")
            return

        # Process each PDF
        all_data = {}

        for pdf_file in pdf_files:
            year = extract_year_from_filename(pdf_file)
            data = parse_annual_report_adaptive(pdf_file, year)

            if data:
                all_data[year] = data
                logger.info(f"  ✓ Extracted {len(data)} fields for {year}")
            else:
                logger.warning(f"  ⚠ No data extracted for {year}")

        if not all_data:
            logger.error("\n[ERROR] No financial data extracted")
            return

        # Create output
        output_file = create_output(all_data)

        logger.info(f"\n{'='*80}")
        logger.info("ANNUAL PARSER COMPLETED")
        logger.info(f"{ '='*80}")
        logger.info(f"✓ Processed {len(pdf_files)} PDF(s)")
        logger.info(f"✓ Extracted data for {len(all_data)} year(s): {list(all_data.keys())}")
        logger.info(f"✓ Output: {output_file}")
        logger.info(f"{ '='*80}")

    except Exception as e:
        logger.error(f"\n[FATAL ERROR] {e}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    main()
